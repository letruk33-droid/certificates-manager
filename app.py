# app.py
import os
import json
import webbrowser
import threading
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from datetime import datetime
import re
import io
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
app.secret_key = 'your_secret_key_here_change_it'

# Файлы для хранения данных
DATA_FILE = 'certificates_data.json'
HISTORY_FILE = 'history.json'

# Новая структура коробок (для добавления и загрузки)
ADD_BOXES = {
    'kommunarka': 'Коммунарка',
    'scherbinka': 'Щербинка',
    'polgoda_tv': 'Полгода т/в',
    'hodili_kommunarka': 'Ходили - Коммунарка',
    'hodili_scherbinka': 'Ходили - Щербинка',
    'nikuda_kommunarka': 'Никуда не ходили - Коммунарка',
    'nikuda_scherbinka': 'Никуда не ходили - Щербинка'
}

# Структура для отображения на главной
MAIN_BOXES = {
    'kommunarka': {'name': 'Коммунарка', 'type': 'single', 'parent': None},
    'scherbinka': {'name': 'Щербинка', 'type': 'single', 'parent': None},
    'polgoda_tv': {'name': 'Полгода т/в', 'type': 'single', 'parent': None},
    'hodili': {
        'name': 'Ходили',
        'type': 'parent',
        'children': {
            'hodili_kommunarka': 'Ходили - Коммунарка',
            'hodili_scherbinka': 'Ходили - Щербинка'
        }
    },
    'nikuda': {
        'name': 'Никуда не ходили',
        'type': 'parent',
        'children': {
            'nikuda_kommunarka': 'Никуда не ходили - Коммунарка',
            'nikuda_scherbinka': 'Никуда не ходили - Щербинка'
        }
    }
}

# Все коробки для поиска
SEARCH_BOXES = {
    'kommunarka': 'Коммунарка',
    'scherbinka': 'Щербинка',
    'polgoda_tv': 'Полгода т/в',
    'hodili_kommunarka': 'Ходили - Коммунарка',
    'hodili_scherbinka': 'Ходили - Щербинка',
    'nikuda_kommunarka': 'Никуда не ходили - Коммунарка',
    'nikuda_scherbinka': 'Никуда не ходили - Щербинка'
}

def load_data():
    """Загрузка данных из файла"""
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_data(data):
    """Сохранение данных в файл"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_history():
    """Загрузка истории"""
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_history(history):
    """Сохранение истории"""
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def add_to_history(action, box_name, fio, series, number_cert, box_number, details=""):
    """Добавление записи в историю"""
    history = load_history()
    history.append({
        'id': datetime.now().timestamp(),
        'datetime': datetime.now().strftime('%d.%m.%Y %H:%M:%S'),
        'action': action,  # 'add', 'give', 'move'
        'box_name': box_name,
        'fio': fio,
        'series': series,
        'number_cert': number_cert,
        'box_number': box_number,
        'details': details
    })
    save_history(history)

def get_free_numbers(box_data):
    """Получение списка свободных номеров в коробке"""
    if not box_data:
        return []
    
    used_numbers = set()
    for cert in box_data.values():
        if isinstance(cert, dict) and 'number' in cert:
            used_numbers.add(cert['number'])
    
    if used_numbers:
        max_number = max(used_numbers)
        free_numbers = [i for i in range(1, max_number + 1) if i not in used_numbers]
    else:
        free_numbers = []
    
    return sorted(free_numbers)

def get_next_new_number(box_data):
    """Получение следующего нового номера"""
    if not box_data:
        return 1
    
    used_numbers = set()
    for cert in box_data.values():
        if isinstance(cert, dict) and 'number' in cert:
            used_numbers.add(cert['number'])
    
    number = 1
    while number in used_numbers:
        number += 1
    
    return number

def add_certificate(box_path, fio, series, number_cert, selected_number=None):
    """Добавление нового сертификата"""
    data = load_data()
    
    # Создаем структуру, если её нет
    if box_path not in data:
        data[box_path] = {}
    
    current = data[box_path]
    
    # Проверяем на дубликаты
    for cert_info in current.values():
        if isinstance(cert_info, dict):
            if cert_info.get('series') == series and cert_info.get('number_cert') == number_cert:
                return False, "Сертификат с такой серией и номером уже существует!"
    
    cert_id = f"{datetime.now().timestamp()}_{fio}"
    
    if selected_number:
        for cert_info in current.values():
            if isinstance(cert_info, dict) and cert_info.get('number') == selected_number:
                return False, f"Номер {selected_number} уже занят!"
        number_to_use = selected_number
    else:
        free_numbers = get_free_numbers(current)
        if free_numbers:
            number_to_use = free_numbers[0]
        else:
            number_to_use = get_next_new_number(current)
    
    current[cert_id] = {
        'fio': fio,
        'series': series,
        'number_cert': number_cert,
        'number': number_to_use,
        'created_at': datetime.now().isoformat(),
        'status': 'active'
    }
    
    save_data(data)
    
    # Добавляем в историю
    box_name = ADD_BOXES.get(box_path, box_path)
    add_to_history('add', box_name, fio, series, number_cert, number_to_use, f"Добавлен сертификат")
    
    return True, f"Сертификат добавлен под номером {number_to_use}"

def remove_certificate(box_path, cert_id):
    """Удаление сертификата (выдача человеку)"""
    data = load_data()
    
    if box_path not in data:
        return False, "Коробка не найдена"
    
    current = data[box_path]
    
    if cert_id not in current:
        return False, "Сертификат не найден"
    
    cert_info = current.pop(cert_id)
    fio = cert_info.get('fio', 'Неизвестно')
    series = cert_info.get('series', '')
    number_cert = cert_info.get('number_cert', '')
    number = cert_info.get('number', '?')
    
    save_data(data)
    
    # Добавляем в историю
    box_name = ADD_BOXES.get(box_path, box_path)
    add_to_history('give', box_name, fio, series, number_cert, number, f"Выдан сертификат")
    
    return True, f"Сертификат {fio} выдан. Номер {number} освобожден"

def delete_all_certificates(box_path=None):
    """Удаление всех сертификатов из коробки или всех коробок"""
    data = load_data()
    
    if box_path:
        if box_path in data:
            count = len(data[box_path])
            # Добавляем записи в историю о массовом удалении
            box_name = ADD_BOXES.get(box_path, box_path)
            for cert_info in data[box_path].values():
                if isinstance(cert_info, dict):
                    add_to_history('give', box_name, cert_info.get('fio', ''), 
                                  cert_info.get('series', ''), cert_info.get('number_cert', ''),
                                  cert_info.get('number', ''), f"Удалено через админку")
            data[box_path] = {}
            save_data(data)
            return True, f"Удалено {count} сертификатов из коробки"
        else:
            return False, "Коробка не найдена"
    else:
        total = 0
        for box_name, box_data in data.items():
            for cert_info in box_data.values():
                if isinstance(cert_info, dict):
                    add_to_history('give', ADD_BOXES.get(box_name, box_name), 
                                  cert_info.get('fio', ''), cert_info.get('series', ''),
                                  cert_info.get('number_cert', ''), cert_info.get('number', ''),
                                  f"Удалено через админку (все коробки)")
                    total += 1
        data = {}
        save_data(data)
        return True, f"Удалено {total} сертификатов из всех коробок"

def search_all_boxes(search_term):
    """Поиск по всем коробкам"""
    data = load_data()
    results = []
    search_lower = search_term.lower()
    
    for box_name, box_data in data.items():
        for cert_id, cert_info in box_data.items():
            if isinstance(cert_info, dict):
                fio = cert_info.get('fio', '')
                series = cert_info.get('series', '')
                number_cert = cert_info.get('number_cert', '')
                
                if (search_lower in fio.lower() or 
                    search_lower in series.lower() or 
                    search_lower in number_cert):
                    
                    box_display_name = ADD_BOXES.get(box_name, box_name)
                    
                    results.append({
                        'id': cert_id,
                        'box': box_display_name,
                        'box_key': box_name,
                        'fio': fio,
                        'series': series,
                        'number_cert': number_cert,
                        'box_number': cert_info.get('number', 0)
                    })
    
    return results

def search_in_box(box_path, search_term):
    """Поиск в конкретной коробке"""
    data = load_data()
    
    if box_path not in data:
        return []
    
    results = []
    search_lower = search_term.lower()
    
    for cert_id, cert_info in data[box_path].items():
        if isinstance(cert_info, dict):
            fio = cert_info.get('fio', '')
            series = cert_info.get('series', '')
            number_cert = cert_info.get('number_cert', '')
            
            if (search_lower in fio.lower() or 
                search_lower in series.lower() or 
                search_lower in number_cert):
                results.append({
                    'id': cert_id,
                    'fio': fio,
                    'series': series,
                    'number_cert': number_cert,
                    'box_number': cert_info.get('number', 0)
                })
    
    return results

def get_box_info(box_path):
    """Получение информации о коробке"""
    data = load_data()
    
    if box_path not in data:
        return [], []
    
    certificates = []
    for cert_id, cert_info in data[box_path].items():
        if isinstance(cert_info, dict):
            certificates.append({
                'id': cert_id,
                'fio': cert_info.get('fio', 'Не указано'),
                'series': cert_info.get('series', 'Не указана'),
                'number_cert': cert_info.get('number_cert', 'Не указан'),
                'box_number': cert_info.get('number', 0)
            })
    
    free_numbers = get_free_numbers(data[box_path])
    return certificates, free_numbers

def parse_certificates_from_text(text):
    """Парсинг сертификатов из текста"""
    lines = text.strip().split('\n')
    certificates = []
    errors = []
    
    for line_num, line in enumerate(lines, 1):
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        
        parts = None
        if '|' in line:
            parts = line.split('|')
        elif ';' in line:
            parts = line.split(';')
        elif '\t' in line:
            parts = line.split('\t')
        
        if not parts or len(parts) < 3:
            errors.append(f"Строка {line_num}: Неверный формат")
            continue
        
        fio = parts[0].strip()
        series = parts[1].strip()
        number_cert = parts[2].strip()
        
        if not re.match(r'N \d \d{8}', number_cert):
            errors.append(f"Строка {line_num}: Неверный формат номера (N 0 00000000)")
            continue
        
        certificates.append({'fio': fio, 'series': series, 'number_cert': number_cert})
    
    return certificates, errors

def add_multiple_certificates(box_path, certificates):
    """Добавление нескольких сертификатов"""
    results = []
    success_count = 0
    
    for cert in certificates:
        success, message = add_certificate(
            box_path, cert['fio'], cert['series'], cert['number_cert'], None
        )
        
        if success:
            success_count += 1
            results.append(f"✅ {cert['fio']} - {message}")
        else:
            results.append(f"❌ {cert['fio']} - {message}")
    
    return success_count, len(certificates) - success_count, results

def export_to_excel(box_filter=None):
    """Экспорт сертификатов в Excel (без pandas)"""
    data = load_data()
    
    # Создаем Excel файл
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем пустой лист
    
    # Если не указана конкретная коробка, не выгружаем ничего
    if not box_filter or box_filter == 'all':
        # Создаем лист с информацией, что нужно выбрать коробку
        ws = wb.create_sheet(title="Информация")
        ws.append(['Для выгрузки выберите конкретную коробку'])
        ws.append(['Перейдите на страницу выгрузки и выберите коробку из списка'])
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FF0000")
        wb.save(output)
        output.seek(0)
        return output
    
    # Проверяем, существует ли коробка
    if box_filter not in data:
        ws = wb.create_sheet(title="Ошибка")
        ws.append(['Коробка не найдена или пуста'])
        wb.save(output)
        output.seek(0)
        return output
    
    box_data = data[box_filter]
    box_name = ADD_BOXES.get(box_filter, box_filter)
    
    # Создаем лист с названием коробки
    sheet_name = box_name[:31]
    ws = wb.create_sheet(title=sheet_name)
    
    # Заголовки (только нужные поля)
    headers = ['Номер в коробке', 'ФИО', 'Серия', 'Номер сертификата']
    ws.append(headers)
    
    # Стиль для заголовков
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Тонкие границы для всех ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Собираем данные
    rows = []
    for cert_id, cert_info in box_data.items():
        if isinstance(cert_info, dict):
            rows.append({
                'number': cert_info.get('number', ''),
                'fio': cert_info.get('fio', ''),
                'series': cert_info.get('series', ''),
                'number_cert': cert_info.get('number_cert', '')
            })
    
    # Сортируем по номеру в коробке
    rows.sort(key=lambda x: x['number'] if isinstance(x['number'], int) else 0)
    
    # Добавляем данные
    for idx, row in enumerate(rows, start=2):
        ws.append([row['number'], row['fio'], row['series'], row['number_cert']])
        
        # Применяем границы к добавленным ячейкам
        for col in range(1, 5):
            cell = ws.cell(row=idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Применяем границы к заголовкам
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Замораживаем заголовок
    ws.freeze_panes = 'A2'
    
    # Добавляем строку с итогами
    total_row = len(rows) + 2
    ws.append([f'ИТОГО: {len(rows)} сертификатов', '', '', ''])
    total_cell = ws.cell(row=total_row, column=1)
    total_cell.font = Font(bold=True)
    total_cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
    
    # Объединяем ячейки для итогов
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    total_cell.alignment = Alignment(horizontal="center")
    
    wb.save(output)
    output.seek(0)
    return output

def move_certificate(from_box, cert_id, to_box, target_number=None):
    """Перемещение сертификата из одной коробки в другую"""
    data = load_data()
    
    # Проверяем существование коробок
    if from_box not in data:
        return False, "Коробка-источник не найдена"
    
    # Создаем целевую коробку, если её нет
    if to_box not in data:
        data[to_box] = {}
    
    from_data = data[from_box]
    to_data = data[to_box]
    
    # Проверяем существование сертификата
    if cert_id not in from_data:
        return False, "Сертификат не найден в исходной коробке"
    
    # Получаем информацию о сертификате
    cert_info = from_data[cert_id]
    fio = cert_info.get('fio', '')
    series = cert_info.get('series', '')
    number_cert = cert_info.get('number_cert', '')
    old_number = cert_info.get('number', '?')
    
    # Определяем новый номер
    if target_number:
        # Проверяем, свободен ли выбранный номер
        for existing_cert in to_data.values():
            if isinstance(existing_cert, dict) and existing_cert.get('number') == target_number:
                return False, f"Номер {target_number} уже занят в целевой коробке!"
        new_number = target_number
    else:
        # Автоматически выбираем свободный или новый номер
        free_numbers = get_free_numbers(to_data)
        if free_numbers:
            new_number = free_numbers[0]
        else:
            new_number = get_next_new_number(to_data)
    
    # Создаем новый сертификат в целевой коробке
    new_cert_id = f"{datetime.now().timestamp()}_{fio}"
    to_data[new_cert_id] = {
        'fio': fio,
        'series': series,
        'number_cert': number_cert,
        'number': new_number,
        'created_at': datetime.now().isoformat(),
        'status': 'active',
        'moved_from': from_box,
        'old_number': old_number,
        'moved_at': datetime.now().isoformat()
    }
    
    # Удаляем из исходной коробки
    del from_data[cert_id]
    
    save_data(data)
    
    # Добавляем в историю
    from_box_name = ADD_BOXES.get(from_box, from_box)
    to_box_name = ADD_BOXES.get(to_box, to_box)
    add_to_history('move', to_box_name, fio, series, number_cert, new_number, 
                   f"Перемещен из {from_box_name} (был №{old_number}) в {to_box_name} (стал №{new_number})")
    
    return True, f"Сертификат {fio} перемещен из {from_box_name} (№{old_number}) в {to_box_name} (№{new_number})"

def get_all_certificates_for_moving(box_path=None):
    """Получение всех сертификатов для перемещения"""
    data = load_data()
    
    if box_path:
        # Получаем сертификаты из конкретной коробки
        if box_path not in data:
            return []
        
        certificates = []
        for cert_id, cert_info in data[box_path].items():
            if isinstance(cert_info, dict):
                certificates.append({
                    'id': cert_id,
                    'fio': cert_info.get('fio', ''),
                    'series': cert_info.get('series', ''),
                    'number_cert': cert_info.get('number_cert', ''),
                    'box_number': cert_info.get('number', 0)
                })
        return sorted(certificates, key=lambda x: x['box_number'])
    else:
        # Получаем все сертификаты из всех коробок
        all_certs = []
        for box_key, box_data in data.items():
            box_name = ADD_BOXES.get(box_key, box_key)
            for cert_id, cert_info in box_data.items():
                if isinstance(cert_info, dict):
                    all_certs.append({
                        'id': cert_id,
                        'box': box_key,
                        'box_name': box_name,
                        'fio': cert_info.get('fio', ''),
                        'series': cert_info.get('series', ''),
                        'number_cert': cert_info.get('number_cert', ''),
                        'box_number': cert_info.get('number', 0)
                    })
        return sorted(all_certs, key=lambda x: (x['box_name'], x['box_number']))

# Декоратор для проверки авторизации
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

# ============== МАРШРУТЫ ==============

@app.route('/')
def index():
    """Главная страница"""
    return render_template('index.html', boxes=MAIN_BOXES)

@app.route('/search')
def search_page():
    """Страница поиска"""
    return render_template('search.html', boxes=SEARCH_BOXES)

@app.route('/add')
def add_page():
    """Страница добавления"""
    return render_template('add.html', boxes=ADD_BOXES)

@app.route('/upload')
def upload_page():
    """Страница загрузки из файла"""
    return render_template('upload.html', boxes=ADD_BOXES)

@app.route('/history')
def history_page():
    """Страница истории"""
    return render_template('history.html')

@app.route('/export')
def export_page():
    """Страница выгрузки Excel"""
    return render_template('export.html', boxes=ADD_BOXES)

@app.route('/move')
def move_page():
    """Страница перемещения сертификатов"""
    return render_template('move.html', boxes=ADD_BOXES)

@app.route('/admin')
@login_required
def admin_page():
    """Админ панель"""
    return render_template('admin.html', boxes=ADD_BOXES)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Вход в админку"""
    if request.method == 'POST':
        password = request.form.get('password')
        if password == 'slavatop':
            session['logged_in'] = True
            return redirect(url_for('admin_page'))
        else:
            return render_template('admin_login.html', error='Неверный пароль')
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    """Выход из админки"""
    session.pop('logged_in', None)
    return redirect(url_for('index'))

@app.route('/api/search-all', methods=['POST'])
def api_search_all():
    """API поиска по всем коробкам"""
    data = request.json
    search_term = data.get('search_term', '')
    
    results = search_all_boxes(search_term)
    return jsonify({'results': results})

@app.route('/api/search-box', methods=['POST'])
def api_search_box():
    """API поиска в конкретной коробке"""
    data = request.json
    box_path = data.get('box_path')
    search_term = data.get('search_term', '')
    
    if not box_path:
        return jsonify({'error': 'Не указана коробка'}), 400
    
    results = search_in_box(box_path, search_term)
    return jsonify({'results': results})

@app.route('/api/add', methods=['POST'])
def api_add():
    """API добавления сертификата"""
    data = request.json
    box_path = data.get('box_path')
    fio = data.get('fio')
    series = data.get('series')
    number_cert = data.get('number_cert')
    selected_number = data.get('selected_number')
    
    if not all([box_path, fio, series, number_cert]):
        return jsonify({'error': 'Заполните все поля'}), 400
    
    if not re.match(r'N \d \d{8}', number_cert):
        return jsonify({'error': 'Номер должен быть в формате: N 0 00000000'}), 400
    
    success, message = add_certificate(box_path, fio, series, number_cert, selected_number)
    
    if success:
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'success': False, 'message': message}), 400

@app.route('/api/upload', methods=['POST'])
def api_upload():
    """API загрузки из файла"""
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не выбран'}), 400
    
    file = request.files['file']
    box_path = request.form.get('box_path')
    
    if not box_path:
        return jsonify({'error': 'Не выбрана коробка'}), 400
    
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400
    
    if not file.filename.endswith('.txt'):
        return jsonify({'error': 'Поддерживаются только файлы .txt'}), 400
    
    try:
        content = file.read().decode('utf-8')
        certificates, parse_errors = parse_certificates_from_text(content)
        
        if parse_errors:
            return jsonify({'success': False, 'parse_errors': parse_errors}), 400
        
        if not certificates:
            return jsonify({'error': 'В файле не найдено корректных сертификатов'}), 400
        
        success_count, error_count, results = add_multiple_certificates(box_path, certificates)
        
        return jsonify({
            'success': True,
            'success_count': success_count,
            'error_count': error_count,
            'results': results,
            'total': len(certificates)
        })
        
    except Exception as e:
        return jsonify({'error': f'Ошибка: {str(e)}'}), 400

@app.route('/api/remove', methods=['POST'])
def api_remove():
    """API удаления сертификата (выдача)"""
    data = request.json
    box_path = data.get('box_path')
    cert_id = data.get('cert_id')
    
    if not all([box_path, cert_id]):
        return jsonify({'error': 'Не указаны поля'}), 400
    
    success, message = remove_certificate(box_path, cert_id)
    
    if success:
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'success': False, 'message': message}), 400

@app.route('/api/box-info/<path:box_path>')
def api_box_info(box_path):
    """Получение информации о коробке"""
    certificates, free_numbers = get_box_info(box_path)
    
    return jsonify({
        'certificates': certificates,
        'free_numbers': free_numbers
    })

@app.route('/api/history')
def api_history():
    """Получение истории"""
    history = load_history()
    return jsonify({'history': history})

@app.route('/api/history/clear', methods=['POST'])
@login_required
def api_history_clear():
    """Очистка истории"""
    save_history([])
    return jsonify({'success': True, 'message': 'История очищена'})

@app.route('/api/export', methods=['POST'])
def api_export():
    """Экспорт в Excel"""
    data = request.json
    box_filter = data.get('box_filter', 'all')
    
    try:
        excel_file = export_to_excel(box_filter)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'certificates_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'error': f'Ошибка при экспорте: {str(e)}'}), 400

@app.route('/api/admin/delete', methods=['POST'])
@login_required
def api_admin_delete():
    """API удаления сертификатов (админка)"""
    data = request.json
    box_path = data.get('box_path')
    
    success, message = delete_all_certificates(box_path if box_path != 'all' else None)
    
    if success:
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'success': False, 'message': message}), 400

@app.route('/api/download-template')
def download_template():
    """Скачать шаблон для заполнения"""
    template_content = """# Шаблон для загрузки сертификатов
# Формат: ФИО | Серия | Номер сертификата
# Номер сертификата: N 0 00000000
# Разделитель: | (вертикальная черта)

Иванов Иван Иванович | к24 | N 1 12345678
Петров Петр Петрович | к24 | N 2 87654321
Сидорова Мария Сергеевна | к24 | N 3 11223344
"""
    
    return send_file(
        io.BytesIO(template_content.encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name='template_certificates.txt'
    )

@app.route('/api/certificates-for-move', methods=['POST'])
def api_certificates_for_move():
    """API получения сертификатов для перемещения"""
    data = request.json
    box_path = data.get('box_path')
    
    certificates = get_all_certificates_for_moving(box_path)
    return jsonify({'certificates': certificates})

@app.route('/api/move', methods=['POST'])
def api_move():
    """API перемещения сертификата"""
    data = request.json
    from_box = data.get('from_box')
    cert_id = data.get('cert_id')
    to_box = data.get('to_box')
    target_number = data.get('target_number')
    
    if not all([from_box, cert_id, to_box]):
        return jsonify({'error': 'Заполните все поля'}), 400
    
    success, message = move_certificate(from_box, cert_id, to_box, target_number)
    
    if success:
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'success': False, 'message': message}), 400

@app.route('/api/box-free-numbers/<path:box_path>')
def api_box_free_numbers(box_path):
    """API получения свободных номеров в коробке"""
    data = load_data()
    
    if box_path not in data:
        return jsonify({'free_numbers': [], 'next_new': 1})
    
    free_numbers = get_free_numbers(data[box_path])
    next_new = get_next_new_number(data[box_path])
    
    return jsonify({
        'free_numbers': free_numbers,
        'next_new': next_new
    })
def open_browser():
    """Открывает браузер через 1 секунду после запуска"""
    webbrowser.open('http://127.0.0.1:5000')
if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
